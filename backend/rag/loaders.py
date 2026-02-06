from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Any

import requests
import yaml

from backend.rag.chunking import chunk_text


@dataclass(frozen=True)
class LoadedDoc:
    title: str
    source: str  # e.g., "toolkit_local" or "web"
    url: str     # for local docs, this will be "local:<filename>"
    text: str


def fetch_url_text(url: str, timeout_s: int = 30) -> str:
    """
    Simple web loader (for public pages). Keeps existing behavior.
    """
    r = requests.get(url, timeout=timeout_s, headers={"User-Agent": "msc-super-friend/1.0"})
    r.raise_for_status()
    return r.text


def load_pdf_text(path: Path) -> str:
    """
    Extracts text from PDF using pypdf. Good enough for MVP.
    """
    try:
        from pypdf import PdfReader
    except Exception as e:
        raise RuntimeError("Missing dependency: pypdf. Install with: pip install pypdf") from e

    reader = PdfReader(str(path))
    chunks: List[str] = []
    for page in reader.pages:
        t = page.extract_text() or ""
        if t.strip():
            chunks.append(t)
    return "\n\n".join(chunks).strip()


def load_xlsx_text(path: Path, max_cells: int = 25000) -> str:
    """
    Extracts text from XLSX by flattening sheets into lines.
    Prevents runaway tokenization via max_cells.
    """
    try:
        import openpyxl
    except Exception as e:
        raise RuntimeError("Missing dependency: openpyxl. Install with: pip install openpyxl") from e

    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    lines: List[str] = []

    seen = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"=== SHEET: {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            row_vals = []
            for v in row:
                if v is None:
                    continue
                s = str(v).strip()
                if not s:
                    continue
                row_vals.append(s)
                seen += 1
                if seen >= max_cells:
                    lines.append("[TRUNCATED: max_cells reached]")
                    return "\n".join(lines).strip()

            if row_vals:
                lines.append(" | ".join(row_vals))

    return "\n".join(lines).strip()


def load_toolkit_local_docs(toolkit_docs_dir: Path) -> List[LoadedDoc]:
    """
    Loads all local docs from backend/data/toolkit_docs.
    Supported: PDF, XLSX
    """
    if not toolkit_docs_dir.exists():
        return []

    docs: List[LoadedDoc] = []
    for p in sorted(toolkit_docs_dir.iterdir()):
        if not p.is_file():
            continue

        ext = p.suffix.lower()
        title = p.stem
        url = f"local:{p.name}"

        if ext == ".pdf":
            text = load_pdf_text(p)
        elif ext in (".xlsx",):
            text = load_xlsx_text(p)
        else:
            # skip unsupported types for MVP
            continue

        if text.strip():
            docs.append(
                LoadedDoc(
                    title=title,
                    source="toolkit_local",
                    url=url,
                    text=text,
                )
            )
    return docs


def load_web_source(url: str, title: str, source: str = "web") -> Optional[LoadedDoc]:
    """
    Loads a web source and returns a LoadedDoc.
    Returns None if loading fails.
    """
    try:
        text = fetch_url_text(url)
        if text.strip():
            return LoadedDoc(
                title=title,
                source=source,
                url=url,
                text=text,
            )
    except Exception:
        return None
    return None


def load_web_sources(sources_path: Path) -> List[Dict[str, Any]]:
    """
    Load web sources from a YAML file.
    Expected structure:
      sources:
        - name: <group_id>
          type: html
          items:
            - title: <title>
              url: <url>
    Returns a list of dicts with: source_id, source_type, title, text, url
    """
    sources_path = Path(sources_path)
    if not sources_path.exists():
        return []

    try:
        raw = yaml.safe_load(sources_path.read_text(encoding="utf-8")) or {}
    except Exception:
        return []

    groups = raw.get("sources") or []
    docs: List[Dict[str, Any]] = []

    for group in groups:
        group_name = str(group.get("name") or "web").strip()
        source_type = str(group.get("type") or "html").strip()
        items = group.get("items") or []

        for i, item in enumerate(items):
            title = str(item.get("title") or item.get("url") or f"{group_name}-{i+1}").strip()
            url = item.get("url")
            if not url:
                continue

            loaded = load_web_source(url=url, title=title, source=group_name)
            if not loaded:
                continue

            docs.append(
                {
                    "source_id": f"{group_name}:{i+1}",
                    "source_type": source_type,
                    "title": loaded.title,
                    "text": loaded.text,
                    "url": loaded.url,
                }
            )

    return docs

def load_sources(sources_path: Path, toolkit_docs_dir: Optional[Path]) -> List[Dict[str, Any]]:
    """
    Unified loader used by ingest.py.
    Combines:
      1) local toolkit docs (PDF/XLSX/etc in toolkit_docs_dir)
      2) web sources listed in sources.yaml (sources_path)
    Returns a list of dicts with at least: { "source": ..., "title": ..., "text": ... }
    """
    docs: List[Dict[str, Any]] = []

    # 1) Local docs
    if toolkit_docs_dir:
        for doc in load_toolkit_local_docs(toolkit_docs_dir):
            docs.append(
                {
                    "source_id": f"{doc.source}:{doc.title}",
                    "source_type": "file",
                    "title": doc.title,
                    "text": doc.text,
                    "url": doc.url,
                    "local_path": str(toolkit_docs_dir / doc.url.replace("local:", "")),
                }
            )

    # 2) Web sources from YAML
    docs.extend(load_web_sources(sources_path))

    return docs
